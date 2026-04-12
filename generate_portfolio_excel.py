from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
# Résumé
ws = wb.active
ws.title = 'Résumé'
rows = [
    ('Champ', 'Détail'),
    ('Nom', 'N’DA YAKPIN YVES JOCELYN'),
    ('Téléphone', '0777828099'),
    ('Email', 'Yvesmryves@gmail.com'),
    ('Localisation', 'Abidjan, Côte d’Ivoire'),
    ('Objectif', 'Jeune professionnel dynamique avec une expérience dans la logistique, la gestion d’équipe et la collecte de données terrain.'),
    ('Profil', 'Organisé, rigoureux, capable de travailler en équipe, maîtrise des outils informatiques de base et bonne capacité d’adaptation.'),
    ('Disponibilité', 'Disponible immédiatement'),
    ('Note', 'Lu et approuvé pour l’étape suivante'),
]
for r in rows:
    ws.append(r)
for cell in ws[1]:
    cell.font = Font(bold=True)

# Compétences
ws = wb.create_sheet('Compétences')
ws.append(('Type', 'Compétence'))
ws.append(('Technique', 'Gestion de stock (notions)'))
ws.append(('Technique', 'Organisation et logistique'))
ws.append(('Technique', 'Travail en équipe'))
ws.append(('Technique', 'Gestion d’équipe'))
ws.append(('Technique', 'Rigueur et sens de l’organisation'))
ws.append(('Technique', 'Collecte et traitement de données'))
ws.append(('Informatique', 'Microsoft Word'))
ws.append(('Informatique', 'Microsoft Excel'))
ws.append(('Informatique', 'PowerPoint'))
ws.append(('Informatique', 'Photoshop'))
ws.append(('Informatique', 'QGIS / QField (bases)'))
ws.append(('Informatique', 'Visual Studio Code (notions)'))
for cell in ws[1]:
    cell.font = Font(bold=True)

# Expériences
ws = wb.create_sheet('Expériences')
ws.append(('Poste', 'Projet', 'Lieu', 'Dates', 'Responsabilités'))
ws.append((
    'Coordinateur terrain',
    'Projet PAVI (MCLU / BNETD)',
    'Daloa, Korhogo, Yamoussoukro',
    '2025',
    'Coordination des équipes terrain; Organisation et suivi des activités; Contrôle qualité des données collectées; Rédaction de rapports; Gestion des tâches et respect des délais'
))
ws.append((
    'Chef d’équipe',
    'Projet d’adressage PADA (MCLU / BNETD)',
    'Abidjan',
    '2022 – 2024',
    'Gestion d’équipe d’agents terrain; Collecte et organisation des données; Participation à l’adressage du District d’Abidjan; Formation des nouveaux agents; Travail en équipe et gestion des zones'
))
ws.append((
    'Supervision installations techniques',
    'Installations caméras & panneaux solaires',
    'Abidjan',
    '2019 – 2020',
    'Suivi des travaux d’installation; Organisation du travail sur site; Coordination avec les techniciens'
))
for cell in ws[1]:
    cell.font = Font(bold=True)

# Formation
ws = wb.create_sheet('Formation')
ws.append(('Diplôme / Certification', 'Année / Niveau', 'Détails'))
ws.append(('Bac A', '2021', ''))
ws.append(('BTS Logistique', 'Niveau en cours', ''))
ws.append(('Formation en infographie', '', ''))
ws.append(('Cisco Intro to Data Science', '', 'Certifié'))
ws.append(('Cisco Intro to Modern AI', '', 'Certifié'))
for cell in ws[1]:
    cell.font = Font(bold=True)

# Fichiers portfolio
ws = wb.create_sheet('Fichiers Portfolio')
ws.append(('Fichier', 'Type', 'Description'))
ws.append(('index.html', 'Page Web', 'Portfolio web existant'))
ws.append(('IMG_6568.jpeg', 'Image', 'Photo ou visuel lié au portfolio'))
ws.append(('IntrotoDataScienceUpdate20260411-32-oui7ni.pdf', 'PDF', 'Certification Cisco Intro to Data Science'))
ws.append(('IntrotoModernAIUpdate20260411-31-u5g7az.pdf', 'PDF', 'Certification Cisco Intro to Modern AI'))
for cell in ws[1]:
    cell.font = Font(bold=True)

wb.save('Yves_Jocelyn_Portfolio.xlsx')
print('Excel file created: Yves_Jocelyn_Portfolio.xlsx')
